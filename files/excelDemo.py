import openpyxl

book = openpyxl.load_workbook("C:/Users/dadu/Downloads/Testing_Udemy.xlsx")
sheet=book.active
# cell = sheet.cell(row=1,column=2)
# print(cell.value)
Dict={}
# add = sheet.cell(row=2, column=2).value="Yejas"
# print(add)
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet['C4'].value)
# for i in range(1,sheet.max_row+1):
#     for j in range(1, sheet.max_column+1):
#         print(sheet.cell(row=i, column=j).value)
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=i,column=j).value

print(Dict)
            # print(sheet.cell(row=i, column=4).value)


